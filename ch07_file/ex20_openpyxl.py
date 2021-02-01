# 파이썬 기초 (+카카오톡 챗봇) 스터디 3주차

# python 파일처리
# 2. 엑셀

"""
OpenPyXL 모듈 설치 후 진행
"""

# 1.1 설치
print("\n[ 1.1 설치 ] ──────────────────────────────────────────────────────────────────────────────────────────────\n")

print('''
* Python 패키지 관리자 PIP 을 이용한 설치
  - 콘솔창에서 pip 명령어로 실행하면 자동으로 설치.

  > pip install openpyxl

* 수동 설치
  - PyCharm 에서는 메뉴에서 직접 라이브러리 설치 가능. File > Settings > Project: [프로젝트명]
    
  - 외부 인터넷이 안되는 환경일 경우
    
  1. OpenPyXL - <https://pypi.org/project/openpyxl/#files> [다운로드]
  2. tar.gz 압축파일을 다운받아서 Python 이 설치된 디렉토리 아래로 이동 후 압축 해제.
     압축파일 안에 setup.py 파일이 존재함.
  3. 콘솔에서 setup.py 파일이 있는 위치로 이동 후, > python setup.py install 명령 실행.
  4. 설치 후 python 을 실행하여 import openpyxl 을 실행해 정상 작동 여부를 확인한다.
  
  ※ install 명령을 수행하기 전에, ```python setup.py build``` 명령을 수행하면 더 좋음.
''')


# 1.2 엑셀 파일 작성
print("\n[ 1.2 엑셀 파일 작성 ] ────────────────────────────────────────────────────────────────────────────────────\n")

import openpyxl as xl

wb = xl.Workbook()

# sheet = wb.create_sheet()  # 워크북은 하나의 sheet 이상으로 구성되기 때문에, 시트를 생성하지 않아도 첫 시트부터 바로 사용 가능.
sheet = wb.active

sheet.title = "성적(원본)"

# 첫 행 (헤더): 컬럼 저장
header_titles = ['이름', '전화전호', '국어', '영어', '수학']
for i, col_name in enumerate(header_titles):
    sheet.cell(row=1, column=i+1, value=col_name)

# 내용 저장
members = [
    ('임요환', '010-1111-1111', 85, 70, 90),
    ('홍진호', '010-2222-2222', 75, 80, 85),
    ('홍진호', '010-2222-2222', 75, 80, 85)
]

for i, member in enumerate(members):
    for col_num, value in enumerate(member):
        sheet.cell(row=2+i, column=col_num+1, value=value)

file_name = "scores.xlsx"
wb.save(f"./sub_dir/{file_name}")
wb.close()

print(f"{file_name}가 sub_dir 폴더에 생성되었습니다.")


# 1.3 엑셀 파일 편집
print("\n[ 1.3 엑셀 파일 편집 ] ────────────────────────────────────────────────────────────────────────────────────\n")

# 엑셀 파일 불러오기
wb = xl.load_workbook("./sub_dir/scores.xlsx", data_only=True)

sheet1 = wb.create_sheet("다음")  # 기본적으로 생성된 시트들 다음의, 마지막 위치에 생성
sheet2 = wb.create_sheet("시트 0", 0)  # 맨 처음 시트 지정 삽입 (기존 시트를 덮어쓰는 건 아님.)
sheet3 = wb.create_sheet("시트 -1", -1)  # 마지막에서 역행하여 시트 삽입

# 시트 삭제
wb.remove(wb["시트 0"])  # 시트명으로만 삭제 가능.(순번으로는 불가능.)
wb.remove(wb["시트 -1"])
wb.remove(wb["다음"])


# 성적(원본) 시트를 복사하여 총점 및 평균을 산출함.
ws = wb.copy_worksheet(wb["성적(원본)"])

ws.title = "성적(산출)"

for data in ws.rows:
    if data[0].row == 1:
        ws.cell(1, 6, "총점")
        ws.cell(1, 7, "평균")
        continue
    else:
        for c, v in enumerate(data):
            print(v.value, end="\n" if c == len(data) - 1 else " ")
        kor = data[2].value
        eng = data[3].value
        mat = data[4].value
        tot = kor + eng + mat
        avg = round((tot / 3), 2)
        ws.cell(row=data[0].row, column=6).value = tot
        ws.cell(row=data[0].row, column=7, value=avg)


# 간단하게 셀 번호로 꺼낼 수 있음.
print(ws["F2"].value)

print([x.value for x in ws["F"]])

wb.save("./sub_dir/scores.xlsx")
wb.close()


# 1.4 다양한 기능
print("\n[ 1.4 다양한 기능 ] ───────────────────────────────────────────────────────────────────────────────────────\n")

wb = xl.load_workbook("./sub_dir/scores.xlsx", data_only=True)

ws = wb["성적(산출)"]

ws.insert_rows(0)
ws.insert_rows(0)

ws.cell(1, 1).value = "성적표"

ws.merge_cells("A1:G1")
ws["A1"].alignment = xl.styles.Alignment(horizontal='center', vertical='center')

wb.save("./sub_dir/scores.xlsx")
wb.close()

# 나머지 생략
